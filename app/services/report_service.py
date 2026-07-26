import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.scoring_service import calculate_results


def generate_excel_report(decision):
    """Generates an Excel binary stream for a decision comparison."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Decision Summary"

    # Styling definitions
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)
    fill_header = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    fill_accent = PatternFill(start_color="F4F4F5", end_color="F4F4F5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Title & Metadata
    ws.append([f"PRISM DECISION REPORT: {decision.title.upper()}"])
    ws.cell(row=1, column=1).font = Font(name="Arial", size=14, bold=True, color="2563EB")
    ws.append([f"Category: {decision.category} | Privacy: {decision.privacy.upper()} | Status: {decision.status_label}"])
    ws.append([f"Goal: {decision.goal or 'N/A'}"])
    ws.append([])

    results = calculate_results(decision)
    if not results or not results['results']:
        ws.append(["No options/criteria configured for evaluation."])
    else:
        # Table Header
        headers = ["Criterion", "Weight", "Priority", "Mandatory"]
        for res in results['results']:
            headers.append(f"{res['option'].name} (Rank #{res['rank']})")
        ws.append(headers)

        header_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center" if col_idx > 4 else "left", vertical="center")

        # Criteria Rows
        score_map = results['score_map']
        for criterion in results['criteria']:
            row_data = [
                criterion.name,
                criterion.weight,
                criterion.priority_label,
                "YES" if criterion.is_mandatory else "NO"
            ]
            for res in results['results']:
                score_obj = score_map.get((criterion.id, res['option'].id))
                raw_val = score_obj.raw_score if score_obj else "—"
                row_data.append(raw_val)
            ws.append(row_data)

            cur_row = ws.max_row
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=cur_row, column=col_idx)
                cell.font = font_regular
                cell.border = thin_border
                if col_idx > 4:
                    cell.alignment = Alignment(horizontal="center")

        # Calculated Score Summary Row
        summary_row = ["Calculated Overall Score (%)", "", "", ""]
        for res in results['results']:
            summary_row.append(f"{res['normalized_score']}%")
        ws.append(summary_row)

        sum_row_idx = ws.max_row
        for col_idx in range(1, len(summary_row) + 1):
            cell = ws.cell(row=sum_row_idx, column=col_idx)
            cell.font = font_bold
            cell.fill = fill_accent
            cell.border = thin_border
            if col_idx > 4:
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf_report(decision):
    """Generates PDF stream using xhtml2pdf or HTML string."""
    from xhtml2pdf import pisa
    from flask import render_template

    results = calculate_results(decision)
    html_content = render_template('reports/pdf_template.html', decision=decision, results=results)

    output = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.StringIO(html_content), dest=output)

    if pisa_status.err:
        return None

    output.seek(0)
    return output
